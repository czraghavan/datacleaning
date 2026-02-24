"""
export.py — Rich Excel export with multiple formatted sheets.

Produces a professional .xlsx with:
  1. Master Data — full output with formatting
  2. Audit Log — mapping decisions and pipeline stats
  3. Vendor Clusters — canonical names and variants
  4. Data Quality — completeness per column with visual bars
"""

import logging
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers,
)
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

logger = logging.getLogger(__name__)

# ── Colour palette ────────────────────────────────────────────────
_HEADER_FILL = PatternFill(start_color="1C1C28", end_color="1C1C28", fill_type="solid")
_HEADER_FONT = Font(name="Inter", bold=True, size=11, color="F0F0F5")
_ALT_ROW_FILL = PatternFill(start_color="F7F7FA", end_color="F7F7FA", fill_type="solid")
_NULL_FILL = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
_GOOD_FILL = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
_WARN_FILL = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
_BAD_FILL = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
_TITLE_FONT = Font(name="Inter", bold=True, size=14, color="6366F1")
_SUBTITLE_FONT = Font(name="Inter", bold=True, size=11, color="333333")
_THIN_BORDER = Border(
    bottom=Side(style="thin", color="E0E0E0"),
)


def export_rich_excel(
    master_df: pd.DataFrame,
    audit_data: dict,
    vendor_clusters: list[dict],
    mapping_log: list[dict],
    job_dir: Path,
    job_id: str,
) -> Path:
    """Create a rich multi-sheet Excel workbook.

    Args:
        master_df: The final processed DataFrame.
        audit_data: Pipeline audit statistics.
        vendor_clusters: List of {canonical, variants} dicts.
        mapping_log: List of {original, canonical, confidence, source} dicts.
        job_dir: Directory to save the file.
        job_id: Unique job identifier.

    Returns:
        Path to the generated .xlsx file.
    """
    wb = Workbook()

    # ── Sheet 1: Master Data ──────────────────────────────────────
    ws_master = wb.active
    ws_master.title = "Master Data"
    _write_master_sheet(ws_master, master_df)

    # ── Sheet 2: Audit Log ────────────────────────────────────────
    ws_audit = wb.create_sheet("Audit Log")
    _write_audit_sheet(ws_audit, audit_data, mapping_log)

    # ── Sheet 3: Vendor Clusters ──────────────────────────────────
    ws_vendors = wb.create_sheet("Vendor Clusters")
    _write_vendor_sheet(ws_vendors, vendor_clusters)

    # ── Sheet 4: Data Quality ─────────────────────────────────────
    ws_quality = wb.create_sheet("Data Quality")
    _write_quality_sheet(ws_quality, master_df)

    # Save
    output_name = f"master_{job_id}.xlsx"
    output_path = job_dir / output_name
    wb.save(str(output_path))

    logger.info("Rich Excel exported: %s (%d rows, 4 sheets)", output_path, len(master_df))
    return output_path


# =====================================================================
# Sheet writers
# =====================================================================

def _write_master_sheet(ws, df: pd.DataFrame):
    """Write the master data with formatting."""
    # Write header
    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _THIN_BORDER

    # Write data
    for row_idx, row in enumerate(df.itertuples(index=False), 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            # Handle NaN/NA
            if pd.isna(value):
                cell.value = None
                cell.fill = _NULL_FILL
            else:
                cell.value = value

            cell.border = _THIN_BORDER

            # Alternate row shading
            if row_idx % 2 == 0 and not pd.isna(value):
                cell.fill = _ALT_ROW_FILL

    # Auto-width columns (capped at 35)
    for col_idx in range(1, len(df.columns) + 1):
        col_letter = get_column_letter(col_idx)
        max_len = len(str(df.columns[col_idx - 1]))
        # Sample first 50 rows for width
        for row_idx in range(2, min(52, len(df) + 2)):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is not None:
                max_len = max(max_len, min(len(str(val)), 35))
        ws.column_dimensions[col_letter].width = max_len + 3

    # Freeze header row
    ws.freeze_panes = "A2"

    # Auto-filter
    if len(df.columns) > 0:
        last_col = get_column_letter(len(df.columns))
        ws.auto_filter.ref = f"A1:{last_col}{len(df) + 1}"


def _write_audit_sheet(ws, audit_data: dict, mapping_log: list[dict]):
    """Write pipeline audit statistics and mapping decisions."""
    row = 1

    # Title
    ws.cell(row=row, column=1, value="Pipeline Audit Report").font = _TITLE_FONT
    row += 2

    # Summary stats
    ws.cell(row=row, column=1, value="Pipeline Statistics").font = _SUBTITLE_FONT
    row += 1

    stats = [
        ("Rows Ingested", audit_data.get("total_rows_ingested", "N/A")),
        ("Tabs Processed", audit_data.get("tabs_processed", "N/A")),
        ("Final Rows", audit_data.get("final_rows", "N/A")),
        ("Unique Vendors", audit_data.get("unique_vendors", "N/A")),
        ("Quote Lines Merged", audit_data.get("quote_lines_merged", "N/A")),
        ("Duplicates Removed", audit_data.get("duplicates_removed", "N/A")),
    ]

    for label, value in stats:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row, column=2, value=value)
        ws.cell(row=row, column=1).border = _THIN_BORDER
        ws.cell(row=row, column=2).border = _THIN_BORDER
        row += 1

    row += 2

    # Column mapping table
    ws.cell(row=row, column=1, value="Column Mapping Decisions").font = _SUBTITLE_FONT
    row += 1

    headers = ["Original Column", "Mapped To", "Confidence", "Source"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.border = _THIN_BORDER
    row += 1

    for mapping in mapping_log:
        ws.cell(row=row, column=1, value=mapping.get("original", "")).border = _THIN_BORDER
        ws.cell(row=row, column=2, value=mapping.get("canonical", "—")).border = _THIN_BORDER

        conf = mapping.get("confidence", "")
        conf_cell = ws.cell(row=row, column=3, value=conf)
        conf_cell.border = _THIN_BORDER
        if conf == "high":
            conf_cell.fill = _GOOD_FILL
        elif conf == "medium":
            conf_cell.fill = _WARN_FILL
        elif conf == "low":
            conf_cell.fill = _BAD_FILL

        ws.cell(row=row, column=4, value=mapping.get("source", "")).border = _THIN_BORDER
        row += 1

    # Unmapped columns
    unmapped = audit_data.get("unmapped_columns", [])
    if unmapped:
        row += 1
        ws.cell(row=row, column=1, value="Unmapped Columns").font = _SUBTITLE_FONT
        row += 1
        for col_name in unmapped:
            ws.cell(row=row, column=1, value=col_name).border = _THIN_BORDER
            ws.cell(row=row, column=1).fill = _BAD_FILL
            row += 1

    # Auto-width
    for col_idx in range(1, 5):
        ws.column_dimensions[get_column_letter(col_idx)].width = 28


def _write_vendor_sheet(ws, vendor_clusters: list[dict]):
    """Write vendor cluster information."""
    row = 1
    ws.cell(row=row, column=1, value="Vendor Clusters").font = _TITLE_FONT
    row += 2

    headers = ["Canonical Name", "Variants", "Variant Count"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.border = _THIN_BORDER
    row += 1

    if not vendor_clusters:
        ws.cell(row=row, column=1, value="No vendor clusters found.").font = Font(italic=True, color="999999")
        row += 1
    else:
        for cluster in vendor_clusters:
            canonical = cluster.get("canonical", "")
            variants = cluster.get("variants", [])
            non_canonical = [v for v in variants if v != canonical]

            ws.cell(row=row, column=1, value=canonical).font = Font(bold=True, color="6366F1")
            ws.cell(row=row, column=1).border = _THIN_BORDER
            ws.cell(row=row, column=2, value=", ".join(non_canonical)).border = _THIN_BORDER
            ws.cell(row=row, column=3, value=len(variants)).border = _THIN_BORDER

            if row % 2 == 0:
                for c in range(1, 4):
                    ws.cell(row=row, column=c).fill = _ALT_ROW_FILL

            row += 1

    # Auto-width
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 15


def _write_quality_sheet(ws, df: pd.DataFrame):
    """Write data quality scorecard: completeness per column."""
    row = 1
    ws.cell(row=row, column=1, value="Data Quality Scorecard").font = _TITLE_FONT
    row += 2

    total_rows = len(df)
    if total_rows == 0:
        ws.cell(row=row, column=1, value="No data to analyze.").font = Font(italic=True, color="999999")
        return

    # Header
    headers = ["Column", "Non-Null Count", "Total Rows", "Completeness %", "Rating"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.border = _THIN_BORDER
    row += 1

    # Compute and sort by completeness
    quality_data = []
    for col in df.columns:
        if col.startswith("meta_"):
            continue
        non_null = int(df[col].notna().sum())
        pct = round(non_null / total_rows * 100, 1)
        quality_data.append((col, non_null, pct))

    quality_data.sort(key=lambda x: x[2], reverse=True)

    for col_name, non_null, pct in quality_data:
        # Rating
        if pct >= 90:
            rating = "Excellent"
            fill = _GOOD_FILL
        elif pct >= 60:
            rating = "Fair"
            fill = _WARN_FILL
        else:
            rating = "Poor"
            fill = _BAD_FILL

        ws.cell(row=row, column=1, value=col_name).border = _THIN_BORDER
        ws.cell(row=row, column=2, value=non_null).border = _THIN_BORDER
        ws.cell(row=row, column=3, value=total_rows).border = _THIN_BORDER

        pct_cell = ws.cell(row=row, column=4, value=pct / 100)
        pct_cell.number_format = numbers.FORMAT_PERCENTAGE
        pct_cell.border = _THIN_BORDER

        rating_cell = ws.cell(row=row, column=5, value=rating)
        rating_cell.fill = fill
        rating_cell.border = _THIN_BORDER
        rating_cell.font = Font(bold=True)

        row += 1

    # Summary row
    row += 1
    overall_non_null = sum(x[1] for x in quality_data)
    overall_total = total_rows * len(quality_data)
    overall_pct = round(overall_non_null / overall_total * 100, 1) if overall_total > 0 else 0

    ws.cell(row=row, column=1, value="OVERALL").font = Font(bold=True, size=12)
    ws.cell(row=row, column=4, value=overall_pct / 100).number_format = numbers.FORMAT_PERCENTAGE
    ws.cell(row=row, column=4).font = Font(bold=True, size=12)

    # Auto-width
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 14
